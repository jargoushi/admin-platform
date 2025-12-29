#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
飞书文档自动截图脚本
使用比特浏览器 + Playwright + GoFullPage 插件实现全页面截图
"""

import asyncio
import json
import os
import re
import time
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict

import httpx
from playwright.async_api import async_playwright, Browser, BrowserContext, Page

# 导入 pyautogui 用于系统级键盘模拟
try:
    import pyautogui
    import pygetwindow as gw
    pyautogui.FAILSAFE = False  # 禁用故障安全（移动鼠标到角落不会中断）
except ImportError:
    print("请先安装 pyautogui pygetwindow: pip install pyautogui pygetwindow")
    exit(1)

# 尝试导入 openpyxl
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    exit(1)


class FeishuScreenshotService:
    """飞书文档截图服务"""

    def __init__(
        self,
        browser_id: str,
        download_dir: str,
        screenshot_output_dir: Optional[str] = None,
        bit_browser_base_url: str = "http://127.0.0.1:54345"
    ):
        """
        初始化截图服务

        Args:
            browser_id: 比特浏览器窗口ID
            download_dir: GoFullPage 截图保存目录（Chrome下载目录）
            screenshot_output_dir: 截图最终保存目录（可选，默认与下载目录相同）
            bit_browser_base_url: 比特浏览器API地址
        """
        self.browser_id = browser_id
        self.download_dir = Path(download_dir)
        self.screenshot_output_dir = Path(screenshot_output_dir) if screenshot_output_dir else self.download_dir
        self.bit_browser_base_url = bit_browser_base_url

        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        self.playwright = None

    async def open_browser(self) -> str:
        """
        通过比特浏览器API打开浏览器，返回 WebSocket 端点

        Returns:
            WebSocket endpoint URL
        """
        url = f"{self.bit_browser_base_url}/browser/open"
        data = {
            "id": self.browser_id,
            "args": [],
            "ignoreDefaultUrls": True
        }

        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(url, json=data)
            response.raise_for_status()
            result = response.json()

            if not result.get("success"):
                raise Exception(f"打开浏览器失败: {result.get('msg', '未知错误')}")

            ws_endpoint = result["data"]["ws"]
            print(f"浏览器已打开，WebSocket: {ws_endpoint}")
            return ws_endpoint

    async def connect_browser(self, ws_endpoint: str):
        """
        通过 Playwright 连接到浏览器

        Args:
            ws_endpoint: WebSocket 端点
        """
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.connect_over_cdp(ws_endpoint)

        # 获取默认上下文
        contexts = self.browser.contexts
        if contexts:
            self.context = contexts[0]
        else:
            self.context = await self.browser.new_context()

        # 获取或创建页面
        pages = self.context.pages
        if pages:
            self.page = pages[0]
        else:
            self.page = await self.context.new_page()

        print("已连接到浏览器")

    async def close_browser(self):
        """关闭浏览器连接"""
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
        print("浏览器连接已关闭")

    async def take_fullpage_screenshot(
        self,
        url: str,
        title: str,
        wait_for_load: int = 5,
        wait_for_screenshot: int = 60
    ) -> Optional[str]:
        """
        对指定URL进行全页面截图（使用 GoFullPage 插件 + PyAutoGUI）

        Args:
            url: 页面URL
            title: 文章标题（用于生成文件名）
            wait_for_load: 页面加载等待时间（秒）
            wait_for_screenshot: 截图完成等待时间（秒）

        Returns:
            截图文件路径，失败返回 None
        """
        try:
            print(f"\n正在处理: {title[:50]}...")

            # 导航到页面
            await self.page.goto(url, wait_until="domcontentloaded", timeout=60000)

            # 等待页面加载
            print(f"  等待页面加载 {wait_for_load} 秒...")
            await asyncio.sleep(wait_for_load)

            # 检查页面标题，判断是否需要登录
            page_title = await self.page.title()
            if "登录" in page_title or "没有权限" in page_title:
                print(f"  ⚠️ 页面需要登录或无权限访问: {page_title}")
                return None

            # 获取截图前的文件列表
            existing_files = set(self.download_dir.glob("*.png"))

            # 激活浏览器窗口到前台
            print("  激活浏览器窗口...")
            await self._activate_browser_window()
            await asyncio.sleep(0.5)

            # 使用 PyAutoGUI 触发 GoFullPage 截图 (Alt+Shift+P)
            print("  触发 GoFullPage 截图 (Alt+Shift+P)...")
            pyautogui.hotkey('alt', 'shift', 'p')

            # 等待截图完成并自动保存
            print(f"  等待截图完成（最长 {wait_for_screenshot} 秒）...")

            saved_path = None
            print(f"  [DEBUG] 监控目录: {self.download_dir}")
            print(f"  [DEBUG] 截图前已有文件数: {len(existing_files)}")
            # 轮询检测新文件，最多等待 wait_for_screenshot 秒
            for i in range(wait_for_screenshot):
                await asyncio.sleep(1)
                current_files = set(self.download_dir.glob("*.png"))
                print(f"  [DEBUG] 第{i+1}秒: 当前文件数={len(current_files)}, 文件={[f.name for f in current_files]}")
                new_files = current_files - existing_files
                if new_files:
                    # 立即复制文件到安全目录，避免被删除
                    latest_file = max(new_files, key=lambda f: f.stat().st_mtime)

                    # 生成安全的文件名
                    safe_title = self._sanitize_filename(title)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    new_filename = f"{safe_title}_{timestamp}.png"

                    # 确保输出目录存在
                    self.screenshot_output_dir.mkdir(parents=True, exist_ok=True)
                    saved_path = self.screenshot_output_dir / new_filename

                    # 使用 shutil.copy2 复制文件（保留原文件）
                    import shutil
                    shutil.copy2(latest_file, saved_path)

                    print(f"  ✅ 截图已保存: {saved_path.name}")
                    break

            # 截图完成后最小化窗口，减少对用户的干扰
            await self._minimize_browser_window()

            if saved_path:
                return str(saved_path)
            else:
                print("  ⚠️ 未检测到新的截图文件")
                return None

        except Exception as e:
            print(f"  ❌ 截图失败: {str(e)}")
            return None

    async def _activate_browser_window(self):
        """激活浏览器窗口到前台"""
        try:
            # 获取页面标题用于查找窗口
            page_title = await self.page.title()

            # 查找包含页面标题的窗口
            windows = gw.getWindowsWithTitle(page_title)
            if windows:
                win = windows[0]
                if win.isMinimized:
                    win.restore()
                win.activate()
                return

            # 如果没找到，尝试查找 Chrome 窗口
            chrome_windows = gw.getWindowsWithTitle("Chrome")
            if not chrome_windows:
                chrome_windows = gw.getWindowsWithTitle("Chromium")

            if chrome_windows:
                win = chrome_windows[0]
                if win.isMinimized:
                    win.restore()
                win.activate()
        except Exception as e:
            print(f"  激活窗口失败: {e}")

    async def _minimize_browser_window(self):
        """最小化浏览器窗口，减少对用户的干扰"""
        try:
            # 获取页面标题用于查找窗口
            page_title = await self.page.title()

            # 查找包含页面标题的窗口
            windows = gw.getWindowsWithTitle(page_title)
            if windows:
                windows[0].minimize()
                return

            # 如果没找到，尝试查找 Chrome 窗口
            chrome_windows = gw.getWindowsWithTitle("Chrome")
            if not chrome_windows:
                chrome_windows = gw.getWindowsWithTitle("Chromium")

            if chrome_windows:
                chrome_windows[0].minimize()
        except Exception as e:
            print(f"  最小化窗口失败: {e}")

    async def _scroll_page(self):
        """滚动页面以加载懒加载内容"""
        try:
            # 获取页面高度
            scroll_height = await self.page.evaluate("document.body.scrollHeight")
            viewport_height = await self.page.evaluate("window.innerHeight")

            # 分段滚动
            current_position = 0
            while current_position < scroll_height:
                current_position += viewport_height
                await self.page.evaluate(f"window.scrollTo(0, {current_position})")
                await asyncio.sleep(0.3)

            # 滚动回顶部
            await self.page.evaluate("window.scrollTo(0, 0)")
            await asyncio.sleep(0.5)
        except Exception as e:
            print(f"  滚动页面时出错: {e}")

    @staticmethod
    def _sanitize_filename(filename: str, max_length: int = 50) -> str:
        """
        清理文件名，移除非法字符

        Args:
            filename: 原始文件名
            max_length: 最大长度

        Returns:
            清理后的文件名
        """
        # 移除非法字符
        illegal_chars = r'[<>:"/\\|?*\n\r\t]'
        sanitized = re.sub(illegal_chars, '', filename)

        # 移除首尾空格
        sanitized = sanitized.strip()

        # 截断长度
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length]

        return sanitized or "untitled"


async def run_screenshot_task(
    json_path: str,
    excel_path: str,
    browser_id: str,
    download_dir: str,
    limit: int = 3
):
    """
    运行截图任务

    Args:
        json_path: 飞书文章JSON文件路径
        excel_path: Excel文件路径
        browser_id: 比特浏览器窗口ID
        download_dir: 下载目录
        limit: 最大处理数量
    """
    # 加载JSON数据
    with open(json_path, 'r', encoding='utf-8') as f:
        articles = json.load(f)

    # 过滤掉无效的文章
    valid_articles = [
        a for a in articles
        if "飞书 - 登录" not in a.get("title", "")
        and "没有权限访问" not in a.get("title", "")
    ]

    # 限制数量
    articles_to_process = valid_articles[:limit]
    print(f"将处理 {len(articles_to_process)} 篇文章")

    # 初始化截图服务
    service = FeishuScreenshotService(
        browser_id=browser_id,
        download_dir=download_dir
    )

    # 存储截图结果
    screenshot_results: List[Dict] = []

    try:
        # 打开并连接浏览器
        ws_endpoint = await service.open_browser()
        await service.connect_browser(ws_endpoint)

        # 给浏览器一些初始化时间
        await asyncio.sleep(2)

        # 逐个处理文章
        for idx, article in enumerate(articles_to_process, start=1):
            title = article.get("title", "")
            url = article.get("url", "")

            print(f"\n[{idx}/{len(articles_to_process)}] 处理中...")

            screenshot_path = await service.take_fullpage_screenshot(
                url=url,
                title=title,
                wait_for_load=5,
                wait_for_screenshot=15
            )

            screenshot_results.append({
                "title": title,
                "url": url,
                "screenshot_path": screenshot_path
            })

            # 文章之间的间隔
            if idx < len(articles_to_process):
                await asyncio.sleep(2)

    finally:
        await service.close_browser()

    # 打印结果汇总
    print("\n" + "=" * 50)
    print("截图任务完成！")
    success_count = sum(1 for r in screenshot_results if r["screenshot_path"])
    print(f"成功: {success_count}/{len(screenshot_results)}")

    # 更新 Excel 文件
    if screenshot_results:
        update_excel_with_screenshots(excel_path, screenshot_results)

    return screenshot_results


def update_excel_with_screenshots(excel_path: str, results: List[Dict]):
    """
    更新 Excel 文件，添加截图路径列

    Args:
        excel_path: Excel文件路径
        results: 截图结果列表
    """
    print(f"\n正在更新 Excel 文件: {excel_path}")

    # 创建 URL 到截图路径的映射
    url_to_screenshot = {r["url"]: r["screenshot_path"] for r in results if r["screenshot_path"]}

    # 加载 Excel
    wb = load_workbook(excel_path)

    # 遍历所有 sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 检查是否已有截图列
        has_screenshot_col = False
        screenshot_col = 4  # 默认第4列

        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "截图路径":
                has_screenshot_col = True
                screenshot_col = col
                break

        # 如果没有截图列，添加表头
        if not has_screenshot_col:
            ws.cell(row=1, column=screenshot_col, value="截图路径")
            ws.cell(row=1, column=screenshot_col).font = Font(bold=True)
            ws.column_dimensions[ws.cell(row=1, column=screenshot_col).column_letter].width = 80

        # 更新数据
        for row in range(2, ws.max_row + 1):
            url = ws.cell(row=row, column=3).value  # URL在第3列
            if url and url in url_to_screenshot:
                ws.cell(row=row, column=screenshot_col, value=url_to_screenshot[url])

    # 保存
    wb.save(excel_path)
    print("Excel 文件已更新！")


async def main():
    """主函数"""
    # 配置参数
    script_dir = Path(__file__).parent
    json_path = script_dir / "feishu_results.json"
    excel_path = script_dir / "feishu_articles_categorized.xlsx"
    browser_id = "e7bc6cca2efd487c8b6dabd70cea8a2d"
    download_dir = r"C:\Users\匠多多\Downloads\13"

    # 运行截图任务（先测试3篇）
    await run_screenshot_task(
        json_path=str(json_path),
        excel_path=str(excel_path),
        browser_id=browser_id,
        download_dir=download_dir,
        limit=5
    )


if __name__ == "__main__":
    asyncio.run(main())
