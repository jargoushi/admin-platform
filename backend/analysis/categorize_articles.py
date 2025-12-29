#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
根据文章标题进行分类，并生成Excel文件，每个分类一个sheet
"""

import json
import re
from pathlib import Path

# 尝试导入 openpyxl，如果没有则提示安装
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    exit(1)


def load_data(json_path: str) -> list:
    """加载JSON数据"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def categorize_article(title: str) -> str:
    """根据标题关键词判断文章分类"""
    title_lower = title.lower()

    # 定义分类规则（优先级从高到低）
    categories = {
        "YouTube视频创作": [
            "youtube", "ytb", "油管", "ypp", "shorts", "订阅",
            "播放量", "频道", "长视频", "动物故事", "印度故事",
            "reddit故事", "猫咪疗愈"
        ],
        "小红书电商": [
            "小红书虚拟", "小红书电商", "小红书店", "虚拟店",
            "虚拟资料", "虚拟电商", "小红书图文", "小红书笔记",
            "小红书自动化", "小红书知识库", "小红书商品"
        ],
        "小红书运营": [
            "小红书", "小绿书", "红薯"
        ],
        "B站好物带货": [
            "b站好物", "b站带货", "b站图书", "b站矩阵",
            "b站深海圈", "b站从0", "b站从零"
        ],
        "B站运营": [
            "b站", "bilibili", "up主", "百万up"
        ],
        "抖音CPS带货": [
            "抖音cps", "抖音自然流", "抖音带货", "直播切片",
            "抖店", "gmv", "cps"
        ],
        "公众号运营": [
            "公众号", "垂直小号", "流量主", "微信公众", "文章二创"
        ],
        "视频号运营": [
            "视频号", "书单", "名人语录"
        ],
        "AI内容创作": [
            "ai短剧", "ai漫剧", "ai视频", "ai内容", "ai创作",
            "ai故事", "ai动画", "ai绘画", "ai短篇", "ai图片",
            "sora", "vidu", "即梦"
        ],
        "AI工具与编程": [
            "cursor", "claude", "gemini", "ai编程", "ai开发",
            "ai工具", "chatgpt", "gpt", "notebooklm", "mcp",
            "figma"
        ],
        "自动化工作流": [
            "n8n", "coze", "扣子", "工作流", "自动化", "rpa",
            "影刀", "飞书多维", "自动发布", "批量"
        ],
        "海外产品SaaS": [
            "saas", "海外", "出海", "美金", "美刀", "刀",
            "工具站", "支付配置", "creem"
        ],
        "闲鱼二手电商": [
            "闲鱼"
        ],
        "快手运营": [
            "快手"
        ],
        "TikTok": [
            "tiktok", "中视频计划"
        ],
        "X-Twitter": [
            "twitter", "推特", "x平台"
        ],
        "个人成长复盘": [
            "复盘", "心路历程", "成长", "逆袭", "裸辞",
            "失业", "离职", "创业", "赚钱", "变现", "闭环",
            "执行力", "坚持", "挣到", "月入", "年薪"
        ]
    }

    # 遍历分类规则进行匹配
    for category, keywords in categories.items():
        for keyword in keywords:
            if keyword in title_lower:
                return category

    # 默认分类
    return "其他"


def create_excel(data: list, output_path: str):
    """创建分类后的Excel文件"""
    # 按分类组织数据
    categorized = {}
    for item in data:
        title = item.get("title", "")
        url = item.get("url", "")

        # 跳过无效数据（如需要登录的页面）
        if "飞书 - 登录" in title or "没有权限访问" in title:
            continue

        category = categorize_article(title)

        if category not in categorized:
            categorized[category] = []
        categorized[category].append({
            "title": title,
            "url": url
        })

    # 创建Excel工作簿
    wb = Workbook()

    # 删除默认sheet
    default_sheet = wb.active

    # 定义样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 按分类创建sheet（sheet名称不能包含特殊字符，如/）
    sheet_order = [
        "YouTube视频创作", "小红书电商", "小红书运营", "B站好物带货", "B站运营",
        "抖音CPS带货", "公众号运营", "视频号运营", "AI内容创作", "AI工具与编程",
        "自动化工作流", "海外产品SaaS", "闲鱼二手电商", "快手运营", "TikTok",
        "X-Twitter", "个人成长复盘", "其他"
    ]

    first_sheet = True
    for category in sheet_order:
        if category not in categorized:
            continue

        articles = categorized[category]

        if first_sheet:
            ws = default_sheet
            ws.title = category[:31]  # Excel sheet名最长31个字符
            first_sheet = False
        else:
            ws = wb.create_sheet(title=category[:31])

        # 添加表头
        ws.cell(row=1, column=1, value="序号").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_alignment

        ws.cell(row=1, column=2, value="标题").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).alignment = header_alignment

        ws.cell(row=1, column=3, value="链接").font = header_font
        ws.cell(row=1, column=3).fill = header_fill
        ws.cell(row=1, column=3).alignment = header_alignment

        # 添加数据
        for idx, article in enumerate(articles, start=1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=article["title"])
            ws.cell(row=idx + 1, column=3, value=article["url"])

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 60

    # 保存文件
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")

    # 输出统计信息
    print("\n=== 分类统计 ===")
    total = 0
    for category in sheet_order:
        if category in categorized:
            count = len(categorized[category])
            total += count
            print(f"{category}: {count}篇")
    print(f"\n总计: {total}篇文章")


def main():
    # 获取当前脚本所在目录
    script_dir = Path(__file__).parent

    # 输入输出路径
    json_path = script_dir / "feishu_results.json"
    output_path = script_dir / "feishu_articles_categorized.xlsx"

    # 加载数据
    data = load_data(str(json_path))
    print(f"共加载 {len(data)} 条数据")

    # 创建Excel
    create_excel(data, str(output_path))


if __name__ == "__main__":
    main()
